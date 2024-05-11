import asyncio
import robin_stocks.robinhood as robin
import os
from openpyxl import Workbook, worksheet, load_workbook
from csv import writer
from datetime import datetime
import sys
from yahooquery import Ticker


def createXLSX(path, file_name=None):
    if file_name == None:
        file_name = 'myDividends.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    row_data = [
    'symbol',
    'record_date',
    'payable_date',
    'quantity',
    'withholding',
    'state',
    'amount',
    'rate'
    ]
    worksheet.append(row_data)
    workbook.save(path + "\\" + file_name)

def clean_sector_data(sector):
    refined_sector = sector.replace("_", " ").lower()
    if refined_sector == 'realestate': # just on example of bad data, could more.
        refined_sector = "real estate"
    return refined_sector

def find_sheet(sheet_name, workbook):
    for worksheet in workbook.worksheets: # change this to seperate function 
        if worksheet.title == sheet_name:
            return worksheet  
    return None    
      
@robin.helper.login_required
def export_dividends(dir_path, file_name=None):

    file_path = dir_path + "\\" + file_name
    workbook = load_workbook(filename=file_path)

    dividend_sheet = find_sheet('Dividends', workbook)

    if dividend_sheet == None:
        dividend_sheet = workbook.create_sheet(title="Dividends")  

    all_dividends = robin.get_dividends()
    
    for row in range(2, dividend_sheet.max_row + 1): # seperate function
        for column in range(1, dividend_sheet.max_column + 1):
            dividend_sheet.cell(row=row, column=column).value = None

    row_index = 2 #start appending at row 2
    for dividend in all_dividends:         
        if(dividend['state'] == 'voided'):  
            continue #don't need voided dividends to be counted
        row_data = [
            robin.get_symbol_by_url(dividend['instrument']),
            datetime.strptime(dividend['record_date'], "%Y-%m-%d"), #as date
            datetime.strptime(dividend['payable_date'], "%Y-%m-%d"), #as date
            float(dividend['position']),
            float(dividend['withholding']),
            dividend['state'],
            float(dividend['amount']),
            float(dividend['rate'])
        ]
        for col, value in enumerate(row_data, start=1):
            dividend_sheet.cell(row=row_index, column=col).value = value
        row_index += 1

    workbook.save(file_path)

@robin.helper.login_required
def export_stocks(dir_path, file_name=None): #find out what is so slow about this function, I believe I need to do calls async
    file_path = dir_path + "\\" + file_name
    workbook = load_workbook(filename=file_path)
    
    stock_sheet = find_sheet("Stock Charts", workbook)
    sector_sheet = find_sheet('Sector Weights', workbook)

    if stock_sheet == None:
        stock_sheet = workbook.create_sheet(title="Stock Charts")  
        
    if sector_sheet == None:
        sector_sheet = workbook.create_sheet(title="Sector Weights") # make these constant variables at the top of the file

    for row in range(2, stock_sheet.max_row + 1): # make this a function
        for column in range(1, stock_sheet.max_column + 1):
            stock_sheet.cell(row=row, column=column).value = None

    for row in range(2, sector_sheet.max_row + 1): # make this a function
        for column in range(1, sector_sheet.max_column + 1):
            sector_sheet.cell(row=row, column=column).value = None

    stocks = robin.get_open_stock_positions()

    row_index = 2 #start at row 2
    portfolio_value = 0

    stocks = robin.get_open_stock_positions()
    stock_tickers = []
    for stock in stocks:
        stock_tickers.append(stock['symbol'])
    latest_prices = robin.get_latest_price(stock_tickers)
    latest_prices_dict = dict(zip(stock_tickers, latest_prices))
    for i in range(len(stocks)):
        portfolio_value += float (latest_prices[i]) * float (stocks[i]['quantity'])
    
    stocks = sorted(stocks, key=lambda x: float (latest_prices_dict[x['symbol']]) * float(x['quantity']), reverse=True)
    sector_totals = {}

    for i in range(len(stocks)):
        total_position = float (latest_prices_dict[stocks[i]['symbol']]) * float (stocks[i]['quantity'])
        row_data = [
            stocks[i]['symbol'],
            float(stocks[i]['quantity']),
            float(stocks[i]['average_buy_price']),
            total_position,
        ]
        for col, value in enumerate(row_data, start=1):
            stock_sheet.cell(row=row_index, column=col).value = value
        row_index += 1
        cur_ticker_symbol = stocks[i]['symbol']
        t = Ticker(cur_ticker_symbol)
        print( "Working on: " + cur_ticker_symbol)
        stock_sector_weights = {}
        if 'sector' in t.asset_profile[cur_ticker_symbol]:
            stock_sector_weights = {t.asset_profile[cur_ticker_symbol]['sector'] : 1}
        else:
            stock_sector_weights = t.fund_sector_weightings.to_dict(orient='dict')[cur_ticker_symbol]
        for sector, weight in stock_sector_weights.items():
            refined_sector = clean_sector_data(sector)
            sector_totals[refined_sector] = sector_totals.get(refined_sector, 0) + total_position * weight

    row_index = 2
    sectors = sorted(sector_totals.items(), key=lambda x: x[1], reverse = True) # sort by value
    for sector, total in sectors:
        sector_sheet.cell(row = row_index, column = 1).value = sector
        sector_sheet.cell(row = row_index, column = 2).value = total
        sector_sheet.cell(row = row_index, column = 3).value = (total / portfolio_value) # Find way to change to percent
        row_index += 1
    workbook.save(file_path)

def export_sectors(dir_path, file_name=None):
    print("not finished")
    
def main():
    args = sys.argv
    file_name = None
    if len(args) > 1:
        file_name =  args[1]
    if file_name == "" or file_name == None:
        file_name = "myDividends.xlsx"

    login = robin.login()

    if not os.path.exists(f"C:\\Users\\{os.getlogin()}\\OneDrive\\Desktop\\" + file_name):
        print("Creating Excel Document")
        createXLSX(f"C:\\Users\\{os.getlogin()}\\OneDrive\\Desktop", file_name)

    #add error protection if file is open
    print("Exporting dividends, please wait. Do not open the excel file until complete.")
    #export_dividends(f"C:\\Users\\{os.getlogin()}\\OneDrive\\Desktop", file_name)
    print("Exporting stock info, please wait. Do not open the excel file until complete.")
    export_stocks(f"C:\\Users\\{os.getlogin()}\\OneDrive\\Desktop", file_name)
    print("running tests")
    print("Portfolio export complete!")


main()
